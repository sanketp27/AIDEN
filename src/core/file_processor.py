"""
File Processor — normalises uploaded files into Gemini-compatible Content parts.

Supported input types:
  Images  : JPEG, PNG, WEBP, GIF, BMP     → inline_data Part (direct vision)
  Audio   : OGG, MP3, WAV, M4A, WEBM     → inline_data Part (Gemini audio)
  PDF     : application/pdf               → inline_data Part (Gemini natively reads PDFs)
  Text    : plain text / CSV / markdown   → text Part
  Office  : DOCX, XLSX, PPTX             → text extraction → text Part
  Fallback: anything else                 → base64 attachment description

The result is a list of `google.genai.types.Part` objects ready to be
appended alongside the user's text Part in a Content message.
"""
from __future__ import annotations

import io
import mimetypes
import structlog
from typing import Optional

log = structlog.get_logger()

# Gemini natively accepts these as inline_data
NATIVE_VISION_MIMES = {
    "image/jpeg", "image/png", "image/webp", "image/gif", "image/bmp",
}
NATIVE_AUDIO_MIMES = {
    "audio/ogg", "audio/mpeg", "audio/mp3", "audio/wav", "audio/x-wav",
    "audio/m4a", "audio/mp4", "audio/webm", "audio/aac",
}
NATIVE_DOC_MIMES = {
    "application/pdf",
    "text/plain", "text/csv", "text/markdown", "text/x-markdown",
}


def detect_mime(filename: str, provided_mime: str) -> str:
    """Return best-guess MIME type, falling back to python mimetypes lib."""
    if provided_mime and provided_mime != "application/octet-stream":
        return provided_mime
    guessed, _ = mimetypes.guess_type(filename)
    return guessed or "application/octet-stream"


async def file_to_parts(
    file_bytes: bytes,
    mime_type:  str,
    filename:   str,
    caption:    str = "",
) -> list:
    """
    Convert raw file bytes into a list of google.genai.types.Part objects.

    Parameters
    ----------
    file_bytes : raw bytes of the uploaded file
    mime_type  : MIME type (may be "application/octet-stream" if unknown)
    filename   : original file name (used for type sniffing & display)
    caption    : optional user caption / instruction for this file

    Returns
    -------
    List of Part objects to append to the Content message.
    """
    from google.genai.types import Part, Blob

    mime_type = detect_mime(filename, mime_type)
    parts: list = []

    if mime_type in NATIVE_VISION_MIMES:
        log.info("file_processor_image", mime=mime_type, size=len(file_bytes))
        parts.append(Part(
            inline_data=Blob(mime_type=mime_type, data=file_bytes)
        ))
        parts.append(Part(text=(
            caption or
            "Analyze this image. Classify its type (whiteboard, document, screenshot, "
            "receipt, business card, photo, etc.), extract all text and actionable content, "
            "then suggest whether I should create tasks, notes, or calendar events from it."
        )))
        return parts

    if mime_type in NATIVE_AUDIO_MIMES:
        log.info("file_processor_audio", mime=mime_type, size=len(file_bytes))
        parts.append(Part(
            inline_data=Blob(mime_type=mime_type, data=file_bytes)
        ))
        parts.append(Part(text=(
            caption or
            "Transcribe this audio recording completely. "
            "Then identify any action items, tasks, decisions, or important information. "
            "Finally, suggest what I should do with this content (create tasks, save a note, etc.)."
        )))
        return parts

    if mime_type == "application/pdf":
        log.info("file_processor_pdf", size=len(file_bytes))
        parts.append(Part(
            inline_data=Blob(mime_type="application/pdf", data=file_bytes)
        ))
        parts.append(Part(text=(
            caption or
            f"This is a PDF file: '{filename}'. "
            "Read and summarize the key content, extract any action items, deadlines, "
            "important decisions, or data. Then suggest what to do with this information."
        )))
        return parts

    if mime_type in NATIVE_DOC_MIMES or mime_type.startswith("text/"):
        try:
            text_content = file_bytes.decode("utf-8", errors="replace")
            # Cap at 50k chars to stay within context window
            if len(text_content) > 50_000:
                text_content = text_content[:50_000] + "\n\n[... truncated at 50,000 characters]"
            log.info("file_processor_text", mime=mime_type, chars=len(text_content))
            parts.append(Part(text=(
                f"File: '{filename}' ({mime_type})\n\n"
                f"{text_content}\n\n"
                f"---\n"
                + (caption or "Analyze this file content. Summarize key information, "
                   "extract action items, and suggest how to use this in AIDEN "
                   "(create tasks, notes, or calendar events).")
            )))
        except Exception as exc:
            log.warning("file_processor_text_decode_failed", error=str(exc))
            parts.append(Part(text=f"File '{filename}' uploaded but could not be decoded as text."))
        return parts


    if (mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or filename.endswith(".docx")):
        text = _extract_docx(file_bytes, filename)
        parts.append(Part(text=text + "\n\n" + (caption or
            "Summarize this document and extract any action items or key decisions.")))
        return parts


    if (mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            or filename.endswith(".xlsx")):
        text = _extract_xlsx(file_bytes, filename)
        parts.append(Part(text=text + "\n\n" + (caption or
            "Analyze this spreadsheet data and summarize key insights.")))
        return parts

    import base64
    b64 = base64.b64encode(file_bytes).decode()[:200]
    log.info("file_processor_fallback", mime=mime_type, filename=filename)
    parts.append(Part(text=(
        f"File attached: '{filename}' (type: {mime_type}, size: {len(file_bytes):,} bytes).\n"
        f"This file type cannot be directly read by the AI. "
        f"Please describe what you'd like to do with it, and I'll assist accordingly.\n"
        + (f"\nYour instruction: {caption}" if caption else "")
    )))
    return parts


def _extract_docx(file_bytes: bytes, filename: str) -> str:
    """Extract plain text from a DOCX file using python-docx."""
    try:
        import docx  # python-docx
        doc = docx.Document(io.BytesIO(file_bytes))
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        text = "\n".join(paragraphs)
        if len(text) > 40_000:
            text = text[:40_000] + "\n[... truncated]"
        log.info("docx_extracted", filename=filename, chars=len(text))
        return f"Document: '{filename}'\n\n{text}"
    except ImportError:
        log.warning("docx_import_failed", hint="pip install python-docx")
        return (f"Document '{filename}' was uploaded but python-docx is not installed. "
                "Please install it to enable DOCX text extraction: pip install python-docx")
    except Exception as exc:
        log.warning("docx_extraction_failed", error=str(exc))
        return f"Document '{filename}' could not be parsed: {exc}"


def _extract_xlsx(file_bytes: bytes, filename: str) -> str:
    """Extract plain text from an XLSX file using openpyxl."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        lines = []
        for sheet_name in wb.sheetnames[:5]:  # max 5 sheets
            ws = wb[sheet_name]
            lines.append(f"\n--- Sheet: {sheet_name} ---")
            for row in ws.iter_rows(max_row=200, values_only=True):  # max 200 rows
                row_vals = [str(c) if c is not None else "" for c in row]
                if any(v.strip() for v in row_vals):
                    lines.append("\t".join(row_vals))
        text = "\n".join(lines)
        if len(text) > 40_000:
            text = text[:40_000] + "\n[... truncated]"
        log.info("xlsx_extracted", filename=filename, chars=len(text))
        return f"Spreadsheet: '{filename}'\n{text}"
    except ImportError:
        log.warning("xlsx_import_failed", hint="pip install openpyxl")
        return (f"Spreadsheet '{filename}' was uploaded but openpyxl is not installed. "
                "Please install it: pip install openpyxl")
    except Exception as exc:
        log.warning("xlsx_extraction_failed", error=str(exc))
        return f"Spreadsheet '{filename}' could not be parsed: {exc}"


def friendly_file_label(mime_type: str, filename: str) -> str:
    """Return a human-readable label for the file type."""
    if mime_type in NATIVE_VISION_MIMES:
        return "Image"
    if mime_type in NATIVE_AUDIO_MIMES:
        return "Audio"
    if mime_type == "application/pdf":
        return "PDF"
    if filename.endswith(".docx"):
        return "Word Document"
    if filename.endswith(".xlsx"):
        return "Spreadsheet"
    if mime_type.startswith("text/"):
        return "Text File"
    return filename.rsplit(".", 1)[-1].upper() if "." in filename else "File"
